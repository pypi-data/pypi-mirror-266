#!/usr/bin/env pythoon3
"""Read AVS dats file."""
import sys
from argparse import ArgumentParser
from pathlib import Path

try:
    import itkdb_gtk
    
except ImportError:
    from pathlib import Path
    cwd = Path(sys.argv[0]).parent.parent
    sys.path.append(cwd.as_posix())

from itkdb_gtk import ITkDBlogin, ITkDButils

import dateutil.parser
import openpyxl as XL
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.exceptions import InvalidFileException

sk_defaults = {
    "institution": "AVS",
    "runNumber": "1",
}


class AVSDataException(Exception):
    """AVSData exception class."""

    def __init__(self, message):
        """Call the base class constructor with the parameters it needs."""
        super().__init__(message)


def create_weight(session, SN, the_date=None, manager="", passed=True, problems=False, comments=[]):
    """Creates the dictionary for a WEIGHT test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        manager: manager name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "WEIGHING", sk_defaults)
    out['component'] = SN
    out['date'] = the_date
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['PRODUCTION_MANAGER'] = manager
    out['comments'] = comments
    return out


def create_manufacturing(session, SN, the_date=None, manager="", passed=True, problems=False, comments=[]):
    """Create the dictionary or the MANUFACTURING test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        manager: manager name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "MANUFACTURING", sk_defaults)
    out['component'] = SN
    out['date'] = the_date
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['PRODUCTION_MANAGER'] = manager
    out['comments'] = comments

    return out


def create_visual_inpection(session, SN, the_date=None, operator="", passed=True, problems=False, comments=[]):
    """Create Visual Inspection test skeleton."""
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "VISUAL_INSPECTION", sk_defaults)
    out['component'] = SN
    out['date'] = the_date
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def create_delamination_test(session, SN, the_date=None, operator="", passed=True, problems=False, comments=[]):
    """Create the delamination test JSON.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        operator: operator name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "DELAMINATION", sk_defaults, {"boolean": False})
    out['component'] = SN
    out['date'] = the_date
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def create_grounding_test(session, SN, the_date=None, operator="", passed=True, problems=False, comments=[]):
    """Create grounding test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        operator: operator name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "GROUNDING_CHECK", sk_defaults, {"boolean": False})
    out['component'] = SN
    out['date'] = the_date
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def create_metrology_test(session, SN, the_date=None, operator="", passed=True, problems=False, comments=[]):
    """Metrology test.

    Args:
        session: the DB session
        SN: Serial Number
        the_date: the date of the test
        operator: operator name
        passed: if test passed or not
        problems: if problems were found during test
        comments: list of comments to append to the test

    """
    the_date = ITkDButils.get_db_date(the_date)
    out = ITkDButils.get_test_skeleton(session, "CORE_PETAL", "METROLOGY_AVS",
                                       sk_defaults, {"integer": -1, "float": -1.0})
    out['component'] = SN
    out['date'] = the_date
    out['passed'] = passed
    out['problems'] = problems
    out['properties']['OPERATOR'] = operator
    out['comments'] = comments

    return out


def split_comp_list(lst):
    """Split a  list of components separated by  various possible characters."""
    if lst is None:
        return []

    if isinstance(lst, float):
        return [lst]

    if isinstance(lst, int):
        return [lst]

    out = [lst]
    for sep in ['/', '\\', '\n']:
        if lst.find(sep) >= 0:
            out = [x.strip() for x in lst.split(sep)]
            break

    return out


def get_comments(txt):
    """Return test DB comment."""
    return split_comp_list(txt)


def get_float(cell, separator=None, default=0.0):
    """Return float from string."""
    txt = cell.value
    if txt is None:
        return default

    if separator is None:
        if isinstance(txt, float):
            return txt

        if isinstance(txt, int):
            return float(txt)

        return default

    else:
        values = []
        for val in split_comp_list(txt):
            if isinstance(val, float) or isinstance(val, int):
                values.append(val)
            else:
                try:
                    v = float(val.strip().replace(',', '.'))
                except ValueError:
                    print("get_float: Cannot convert {} in {}".format(val, cell.coordinate))
                    v = default

                values.append(v)

        return values


def get_boolean(cell):
    """Get a boolean from a cell."""
    value = cell.value
    if value is None:
        return False

    else:
        txt = value.strip().lower()
        return (txt == "pass")


def get_text(cell):
    """Get a string from a cell."""
    value = cell.value
    if value:
        value = value.strip()

    return value


def get_res_and_accep(sheet, indx):
    """Return result and acceptancee."""
    val = get_float(sheet["g{}".format(indx)])
    accept = get_boolean(sheet["h{}".format(indx)])
    return val, accept


def find_idx(lst, val):
    """Return index of occurence of val in lst."""
    idx = [i for i, x in enumerate(lst) if (x and x.find(val) >= 0)]
    return idx


def cell_value(sheet, coord):
    """Return the cell value."""
    cell = sheet[coord]
    if not isinstance(cell, MergedCell):
        return cell.value

    # "Oh no, the cell is merged!"
    for range in sheet.merged_cells.ranges:
        if coord in range:
            return range.start_cell.value

    return None


def check_for_problems(sheet, the_test, row_range):
    """Finds FAIL massages in the Acceptance column."""
    nfail = 0
    for row in range(row_range[0], row_range[1]):
        txt = get_text(sheet["h{}".format(row)])
        if txt is None:
            continue

        txt = txt.lower()
        if txt[0] == 'f':
            nfail += 1
            hdr = get_text(sheet["d{}".format(row)])
            reason = cell_value(sheet, "i{}".format(row))

            if len(reason) < 20:
                msg = "{}: {}".format(hdr, reason)
                the_test["defects"].append({"name": msg})
            else:
                the_test["defects"].append({"name": hdr,
                                            "description": reason})

    if nfail:
        the_test["passed"] = False


def get_coreID(bustapeID):
    """Build core SN from bus tape SN."""
    SN = "20USEBC" + bustapeID[-7:]
    return SN


def readFATfile(session, fnam, SN=None):
    """Read data from FAT excel file.

    Args:
        session: the DB session
        fnam: File path
        SN: COre serial number

    """
    # Open spreadsheet
    try:
        wb = XL.load_workbook(fnam)
    except InvalidFileException as ee:
        print("Could not open input file: ", fnam)
        print(ee)
        return None

    # Assume active sheet is the good one, otherwise will have to find in wb.sheetnames
    sheet = wb.active
    if sheet.max_row < 50 or sheet.max_column < 9:
        raise AVSDataException("Wrong FAT file")

    # Check the SN of the petal core
    if SN is None or len(SN) == 0:
        coreID = split_comp_list(sheet['C6'].value)
        if len(coreID) == 0:
            raise AVSDataException("Cannot figure out core SN in FAT file.")

        for cID in coreID:
            cmp = ITkDButils.get_DB_component(session, cID)
            if cmp["type"]["code"] == "BT_PETAL_FRONT":
                SN = get_coreID(cID)
                break

    batch = sheet['E6'].value
    operator = sheet['G6'].value

    txt = list(map(str.strip, sheet['i4'].value.split(':')))[1]
    test_date = dateutil.parser.parse(txt)

    # Get the test index
    test_name = [str(sheet[x][1].value) for x in range(1, sheet.max_row)]
    tests = [str(sheet[x][3].value) for x in range(1, sheet.max_row)]

    # This is to avoid adding 1 for cell names...
    tests.insert(0, None)
    test_name.insert(0, None)

    #
    # Visual inspection
    vi_text = get_text(sheet['i9'])
    vi_result = get_text(sheet['g9'])
    vi_pass = (sheet['h9'].value.strip().lower() == "pass")
    vi_defects = []
    if vi_pass:
        if vi_result and len(vi_result):
            if vi_text and len(vi_text):
                vi_text = vi_result + '\n' + vi_text
            else:
                vi_text = vi_result

    else:
        vi_defects.append({"name": "PETAL_VI_DEFECT", "description": vi_result})

    vi_test = create_visual_inpection(session, SN, test_date, operator, vi_pass,
                                      comments=get_comments(vi_text))
    for df in vi_defects:
        vi_test["defects"].append(df)

    #
    # Delamination test
    dl_text = get_text(sheet['i10'])
    dl_result = get_text(sheet['g10'])
    dl_pass = (sheet['h10'].value.strip().lower() == "pass")
    dl_defects = []
    if dl_pass:
        if dl_result and len(dl_result):
            if dl_text and len(dl_text):
                dl_text = dl_result + '\n' + dl_text
            else:
                dl_text = dl_result

    else:
        dl_defects.append({"name": "PETAL_DL_DEFECT",
                           "description": dl_result})

    delamination_test = create_delamination_test(session, SN, test_date, operator, dl_pass,
                                                 comments=get_comments(dl_text))
    for df in dl_defects:
        delamination_test["defects"].append(df)

    #
    # Conductivity
    # TODO: read proper rows
    grounding_test = create_grounding_test(session, SN, test_date, operator)
    cond_val, cond_pass = get_res_and_accep(sheet, tests.index("COND"))
    if "INS_LOOP" in tests:
        loop_val, loop_pass = get_res_and_accep(sheet, tests.index("INS_LOOP"))
    else:
        loop_val, loop_pass = get_res_and_accep(sheet, tests.index("INS"))

    if "INS_LOOP_GND" in tests:
        loop_gnd_val, loop_gnd_pass = get_res_and_accep(sheet, tests.index("INS_LOOP_GND"))
    else:
        loop_gnd_val, loop_gnd_pass = get_res_and_accep(sheet, tests.index("INS_FACE"))

    grounding_test["results"]["RESISTANCE_FB"] = cond_val
    grounding_test["results"]["RESISTANCE_PIPES"] = loop_val
    grounding_test["results"]["RESISTANCE_PIPE_GND"] = loop_gnd_val
    check_for_problems(sheet, grounding_test, [tests.index('COND'), tests.index("WEIGH")])

    #
    # Weight
    petal_weight, weight_pass = get_res_and_accep(sheet, tests.index("WEIGH"))

    #
    # Metrology AVS
    metrology_test = create_metrology_test(session, SN, test_date, operator)
    metrology_test["results"]["LOCATOR1_DIAMETER"] = get_float(sheet['g{}'.format(tests.index("PL01_DIAM"))])
    metrology_test["results"]["LOCATOR2_DIAMETER"] = get_float(sheet['g{}'.format(tests.index("PL02_DIAM"))])
    metrology_test["results"]["LOCATOR3_DIAMETER"] = get_float(sheet['g{}'.format(tests.index("PL03_DIAM"))])
    metrology_test["results"]["LOCATOR2_X"] = get_float(sheet['g{}'.format(find_idx(tests, "PL02_X")[0])])
    metrology_test["results"]["LOCATOR2_Y"] = get_float(sheet['g{}'.format(find_idx(tests, "PL02_Y")[0])])
    metrology_test["results"]["LOCATOR3_X"] = get_float(sheet['g{}'.format(find_idx(tests, "PL03_X")[0])])
    metrology_test["results"]["LOCATOR3_Y"] = get_float(sheet['g{}'.format(find_idx(tests, "PL03_Y")[0])])
    metrology_test["results"]["FIDUCIAL1_DIAMETER"] = get_float(sheet["g{}".format(find_idx(tests, "FD01_DIAM")[0])])
    metrology_test["results"]["FIDUCIAL1_X"] = get_float(sheet["g{}".format(find_idx(tests, "FD01_X")[0])])
    metrology_test["results"]["FIDUCIAL1_Y"] = get_float(sheet["g{}".format(find_idx(tests, "FD01_Y")[0])])
    metrology_test["results"]["FIDUCIAL2_DIAMETER"] = get_float(sheet["g{}".format(find_idx(tests, "FD02_DIAM")[0])])
    metrology_test["results"]["FIDUCIAL2_X"] = get_float(sheet["g{}".format(find_idx(tests, "FD02_X")[0])])
    metrology_test["results"]["FIDUCIAL2_Y"] = get_float(sheet["g{}".format(find_idx(tests, "FD02_Y")[0])])
    metrology_test["results"]["ANGLE_VCHANNEL"] = get_float(sheet["g{}".format(find_idx(tests, "VANGL")[0])])
    metrology_test["results"]["ENVELOPE"] = get_float(sheet["g{}".format(find_idx(tests, "ENVEL")[0])])
    metrology_test["results"]["COPLANARITY_FRONT"] = get_float(sheet["g{}".format(find_idx(tests, "F/PL_PLAN")[0])])
    metrology_test["results"]["LOCAL_FLATNESS_FRONT"] = get_float(sheet["g{}".format(find_idx(tests, "F/FS_PLAN")[0])], '/')
    metrology_test["results"]["PARALLELISM_FRONT"] = get_float(sheet["g{}".format(find_idx(tests, "F/PARAL")[0])])
    metrology_test["results"]["COPLANARITY_BACK"] = get_float(sheet["g{}".format(find_idx(tests, "B/PL_PLAN")[0])])
    metrology_test["results"]["LOCAL_FLATNESS_BACK"] = get_float(sheet["g{}".format(find_idx(tests, "B/FS_PLAN")[0])], '/')
    metrology_test["results"]["PARALLELISM_BACK"] = get_float(sheet["g{}".format(find_idx(tests, "B/PARAL")[0])])

    # Get defects
    check_for_problems(sheet, metrology_test, [tests.index("WEIGH")+1, sheet.max_row])

    return vi_test, delamination_test, grounding_test, metrology_test, batch, petal_weight


def readProductionSheet(session, fnam, SN):
    """Read data fro AVS PS.

    Args:
        session: the DB session
        fnam: path of input file
        SN: The serial number
        write_json: if true, test json is writen to file.

    """
    try:
        wb = XL.load_workbook(fnam)
    except InvalidFileException as ee:
        print("Could not open input file: ", fnam)
        print(ee.message)
        return None

    # Assume active sheet is the good one, otherwise will have tofind in wb.sheetnames
    sheet = wb.active
    if sheet.max_row > 30 or sheet.max_column > 8:
        raise AVSDataException("Wrong PS file")

    ID = sheet['c12'].value.strip()
    SN = get_coreID(ID)
    mould_id = sheet['b9'].value
    txt = list(map(str.strip, sheet['f4'].value.split(':')))[1]
    try:
        test_date = dateutil.parser.parse(txt)
    except Exception:
        test_date = ITkDButils.get_db_date(sheet['e8'].value)

    manager = sheet['e7'].value

    # Manufacturing
    comments = get_comments(sheet['a25'].value)
    manufacturing = create_manufacturing(session, SN, test_date, manager, comments=comments)
    manufacturing['properties']['START_DATE'] = ITkDButils.get_db_date(sheet['e8'].value)
    manufacturing['properties']['FINISH_DATE'] = ITkDButils.get_db_date(sheet['e9'].value)
    manufacturing["properties"]["MOULD_ID"] = mould_id
    manufacturing["properties"]["PROCESS_DOCUMENT"] = sheet['e6'].value
    manufacturing["results"]["LOCATOR_A"] = sheet['c14'].value
    manufacturing["results"]["LOCATOR_B"] = sheet['c15'].value
    manufacturing["results"]["LOCATOR_C"] = sheet['c16'].value
    manufacturing["results"]["HONEYCOMBSET"] = split_comp_list(sheet['c17'].value)
    manufacturing["results"]["EPOXY_ADHESIVE"] = split_comp_list(sheet['c20'].value)
    manufacturing["results"]["EPOXY_PUTTY"] = split_comp_list(sheet['c21'].value)
    manufacturing["results"]["EPOXY_CONDUCTIVE"] = split_comp_list(sheet['c22'].value)

    # Weighing
    weighing = create_weight(session, SN, test_date, manager)
    comp_weight = [get_float(x[0]) for x in sheet['d12:d22']]
    petal_weight = sum([float(x) for x in comp_weight])
    weighing["results"]["WEIGHT_FACING_FRONT"] = comp_weight[0]
    weighing["results"]["WEIGHT_FACING_BACK"] = comp_weight[1]
    weighing["results"]["WEIGHT_LOCATOR_A"] = comp_weight[2]
    weighing["results"]["WEIGHT_LOCATOR_B"] = comp_weight[3]
    weighing["results"]["WEIGHT_LOCATOR_C"] = comp_weight[4]
    weighing["results"]["WEIGHT_COOLINGLOOPASSEMBLY"] = comp_weight[6]
    weighing["results"]["WEIGHT_HONEYCOMBSET"] = comp_weight[5]
    weighing["results"]["WEIGHT_EPOXYADHESIVE"] = comp_weight[8]
    weighing["results"]["WEIGHT_EPOXYPUTTY"] = comp_weight[9]
    weighing["results"]["WEIGHT_EPOXYCONDUCTIVE"] = comp_weight[10]
    weighing["results"]["WEIGHT_CORE"] = petal_weight

    # Comments
    for i in range(12, 23):
        cell_id = sheet['B{}'.format(i)].value
        comment = sheet['E{}'.format(i)].value
        if comment is not None:
            comment = comment.strip()
            msg = "{}: {}".format(cell_id, comment)
            weighing["comments"].append(msg)

    DESY = {
        "FacingFront": sheet['c12'].value.strip(),
        "FacingBack": sheet['c13'].value.strip(),
        "CoolingLoop": str(sheet['c18'].value),
        "AllcompSet": sheet['c19'].value,
        "HoneyCombSet": manufacturing["results"]["HONEYCOMBSET"]
    }

    return manufacturing, weighing, DESY


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument('files', nargs='*', help="Input files")
    parser.add_argument("--SN", dest="SN", type=str, default="SNnnn",
                        help="Module serial number")

    options = parser.parse_args()
    if len(options.files) == 0:
        print("I need an input file")
        sys.exit()

    dlg = ITkDBlogin.ITkDBlogin()
    client = dlg.get_client()
    if client is None:
        print("Could not connect to DB with provided credentials.")
        dlg.die()
        sys.exit()

    fnam = Path(options.files[0]).expanduser().resolve()
#    readProductionSheet(fnam, options.SN)
    readFATfile(client, fnam, options.SN)
    dlg.die()
