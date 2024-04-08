import contextlib
from functools import wraps

from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


@contextlib.contextmanager
def load_workbook(filepath):
    '''Context manager for openpyxl workbooks.'''

    workbook = openpyxl_load_workbook(filepath)

    try:
        yield workbook
    except Exception as e:
        print(e)
    finally:
        workbook.save(filepath)
        workbook.close()


class XLFormatter:
    """Formatter for Excel files that utilizes the OpenPyXl engine."""

    def requires_sheet(self, func):
        @wraps(func)
        def wrapper(self, *args, **kwargs):
            if self._sheet is None:
                raise AttributeError(
                    "Use 'XLFormatter.set_sheet()' to select a sheet")
            return func(self, *args, **kwargs)
        return wrapper

    def __init__(self, workbook, sheet=None):
        self._workbook = workbook
        self._sheet = sheet

        if self._workbook is None:
            raise AttributeError(
                "XLFormatter requires a 'workbook' parameter that wasn't given")

    def set_sheet(self, sheet):
        """Changes the sheet."""

        if isinstance(sheet, str):
            self._sheet = self._workbook[sheet]
        else:
            self._sheet = sheet
        return self

    @requires_sheet
    def align_all(self, column_names, **kwargs) -> None:
        """Aligns all used cells in the sheet."""

        default_keyword_args = {
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": True,
            "indent": 1
        }

        align_kwargs = kwargs.pop('align_settings', default_keyword_args)
        header_align_kwargs = kwargs.pop('header_align_settings', align_kwargs)

        for col_index in column_names:
            for row_index, cell in enumerate(self._sheet[get_column_letter(col_index + 1)]):
                cell.alignment = Alignment(
                    **(header_align_kwargs if row_index == 0 else align_kwargs))

        return self

    @requires_sheet
    def autofit_columns(self):
        """Autofit all columns in the sheet"""

        for column_cells in self._sheet.columns:
            length = max(len(str(cell.value or ""))
                         for cell in column_cells)
            self._sheet.column_dimensions[column_cells[0]
                                          .column_letter].width = length + 5

        return self
