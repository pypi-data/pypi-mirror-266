from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ficamp.datastructures import Currency, Tx
from ficamp.parsers.protocols import ParserProtocol


class AccountBBVAParser(ParserProtocol):
    """Parser for BBVA bank account extract"""

    def load(self, filename: Path):
        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 6
        start_column = 2

        self.rows = [
            row
            for row in sheet.iter_rows(
                min_row=start_row, min_col=start_column, values_only=True
            )
        ]

    def parse(self) -> list[Tx]:
        return [
            self.row_processor(row)
            for row in self.rows
            if self.row_processor(row) is not None
        ]

    def row_processor(self, row):
        # Skip Credit Card charge in Account
        if "targeta" in row[2] or "tarjeta" in row[2]:
            return None

        concept = f"{row[2]} || {row[3]}"

        return Tx(
            date=row[0],
            amount=Decimal(str(row[4])),
            currency=Currency(row[5]),
            concept=concept,
            category=None,
            tx_metadata={"more_details": row[8], "origin": "BBVA Account"},
            tags=[],
        )


class CreditCardBBVAParser(ParserProtocol):
    """Parser for BBVA Credit Card Extract"""

    def load(self, filename: Path):
        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 6
        start_column = 2

        self.rows = [
            row
            for row in sheet.iter_rows(
                min_row=start_row, min_col=start_column, values_only=True
            )
        ]

    def parse(self) -> list[Tx]:
        return [
            self.row_processor(row)
            for row in self.rows
            if self.row_processor(row) is not None
        ]

    def row_processor(self, row):
        # Skip a positive record, as it's an internal operation for the bank
        # to "recharge" the credit card when it gets discounted from associated account
        if row[3] > 0:
            return None

        return Tx(
            date=datetime.strptime(row[0], "%d/%m/%Y"),
            amount=Decimal(str(row[3])),
            currency=Currency("EUR"),
            concept=row[2],
            category=None,
            tx_metadata={"origin": "BBVA Credit Card"},
            tags=[],
        )
