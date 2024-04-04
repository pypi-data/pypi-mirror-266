from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from ficamp.datastructures import Currency, Tx
from ficamp.parsers.protocols import ParserProtocol


class CaixaBankParser(ParserProtocol):
    """Parser for CaixaBank bank account extract"""

    def load(self, filename: Path):
        wb = load_workbook(filename)
        sheet = wb.active
        start_row = 4
        start_column = 1

        self.rows = [
            row
            for row in sheet.iter_rows(
                min_row=start_row, min_col=start_column, values_only=True
            )
        ]

    def parse(self) -> list[Tx]:
        return [
            self.row_processor(row)
            for row in self.rows
            if self.row_processor(row) is not None
        ]

    def row_processor(self, row):
        concept = f"{row[2]} || {row[3]}"

        return Tx(
            date=row[0],
            amount=Decimal(str(row[4])),
            currency=Currency("EUR"),
            concept=concept,
            category=None,
            tx_metadata={"origin": "CaixaBank Account"},
            tags=[],
        )
