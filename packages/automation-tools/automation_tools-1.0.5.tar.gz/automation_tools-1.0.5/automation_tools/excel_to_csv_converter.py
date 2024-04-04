# excel_to_csv_converter.py
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import threading
import openpyxl

class ExcelToCsvConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel to CSV Converter")

        self.loaded_files = []
        self.output_directory = ""
        self.csv_file_count = 0  # To track the count of CSV files

        self.file_list_label = tk.Label(self.root, text="Loaded Files:")
        self.file_list_label.pack(padx=10, pady=5)

        self.file_listbox = tk.Listbox(self.root, selectmode=tk.MULTIPLE, width=40, height=5)
        self.file_listbox.pack(padx=10, pady=5)

        self.load_files_button = tk.Button(self.root, text="Load XLSX Files", command=self.load_xlsx_files)
        self.load_files_button.pack(padx=10, pady=5)

        self.output_dir_label = tk.Label(self.root, text="Output Directory:")
        self.output_dir_label.pack(padx=10, pady=5)

        self.output_dir_text = tk.Label(self.root, text=self.output_directory)
        self.output_dir_text.pack(padx=10, pady=5)

        self.output_dir_button = tk.Button(self.root, text="Select Output Directory", command=self.select_output_directory)
        self.output_dir_button.pack(padx=10, pady=5)

        self.csv_count_label = tk.Label(self.root, text="CSV Files Added:")
        self.csv_count_label.pack(padx=10, pady=5)

        self.csv_count_text = tk.Label(self.root, text="0")
        self.csv_count_text.pack(padx=10, pady=5)

        self.convert_button = tk.Button(self.root, text="Convert to CSV", command=self.convert_to_csv)
        self.convert_button.pack(padx=10, pady=10)

        self.progress_label = tk.Label(self.root, text="Progress:")
        self.progress_label.pack(padx=10, pady=5)

        self.progress_var = tk.StringVar()
        self.progress_text = tk.Label(self.root, textvariable=self.progress_var)
        self.progress_text.pack(padx=10, pady=5)

        self.sheet_names_label = tk.Label(self.root, text="Sheet Names:")
        self.sheet_names_label.pack(padx=10, pady=5)

        self.sheet_names_text = tk.Text(self.root, height=5, width=40)
        self.sheet_names_text.pack(padx=10, pady=5)

    def load_xlsx_files(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
        for file_path in file_paths:
            self.loaded_files.append(file_path)
            self.file_listbox.insert(tk.END, os.path.basename(file_path))

    def select_output_directory(self):
        self.output_directory = filedialog.askdirectory()
        self.output_dir_text.config(text=self.output_directory)  # Update the label text

    def convert_to_csv(self):
        if not self.loaded_files:
            messagebox.showerror("Error", "No files loaded.")
            return

        if not self.output_directory:
            messagebox.showerror("Error", "No output directory selected.")
            return

        self.convert_thread = threading.Thread(target=self.convert_and_update_progress)
        self.convert_thread.start()

    def convert_and_update_progress(self):
        total_files = len(self.loaded_files)
        progress_step = 100 / total_files
        current_progress = 0

        for idx, file_path in enumerate(self.loaded_files):
            try:
                wb = openpyxl.load_workbook(file_path)
                file_name = os.path.splitext(os.path.basename(file_path))[0]
                tabs = wb.sheetnames

                sheet_names_str = ", ".join(tabs)
                self.sheet_names_text.delete(1.0, tk.END)
                self.sheet_names_text.insert(tk.END, sheet_names_str)

                for tab in tabs:
                    ws = wb[tab]
                    csv_path = os.path.join(self.output_directory, f"{file_name}_{tab}.csv")
                    with open(csv_path, "w") as f:
                        for row in ws.iter_rows():
                            row_data = [str(cell.value) for cell in row]
                            f.write(",".join(row_data) + "\n")
                    self.csv_file_count += 1  # Increment the count of CSV files
                    self.csv_count_text.config(text=str(self.csv_file_count))  # Update the count label

                current_progress += progress_step
                self.progress_var.set(f"{idx+1}/{total_files} files converted")
                self.root.update()

            except Exception as e:
                messagebox.showerror("Error", f"Error converting file: {file_path}\n{e}")
                return

        messagebox.showinfo("Conversion Complete", "All files converted successfully.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToCsvConverter(root)
    root.mainloop()
